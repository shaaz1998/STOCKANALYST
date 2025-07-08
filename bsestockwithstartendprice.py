import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 🔁 Set your tickers here
tickers = {
    "LINC": "LINC.BO",
    "QUICKHEAL": "QUICKHEAL.BO",
    "RAYMOND": "RAYMOND.BO",
    "TCS": "TCS.BO",
    "HCLTECH": "HCLTECH.BO",
    "BBTC": "BBTC.BO",
    "EIDPARRY": "EIDPARRY.BO",
    "HINDALCO": "HINDALCO.BO",
    "DEVYANI": "DEVYANI.BO",
    "AJANTPHARM": "AJANTPHARM.BO",
    "SAAKSHI-ST": "SAAKSHI.NS",
    "SABEVENTS": "SABEVENTS.BO",
    "SABOOSOD": "SABOOSOD.BO",
    "MUFIN": "MUFIN.NS",
    "BEARDSELL": "BEARDSELL.NS"

}

years = range(2020, 2025)
rows = []

for name, symbol in tickers.items():
    row = {"STOCK": name}
    for year in years:
        start_date = f"{year}-07-01"
        end_date = f"{year}-08-30"
        try:
            data = yf.download(symbol, start=start_date, end=end_date, auto_adjust=True, progress=False)

            if data.empty:
                raise ValueError("No data found.")

            start_data = data.loc[data.index >= start_date]
            end_data = data.loc[data.index <= end_date]

            if start_data.empty or end_data.empty:
                raise ValueError("No trading data in date range.")

            start_row = start_data.iloc[0]
            end_row = end_data.iloc[-1]

            start_price = round(float(start_row['Close'].iloc[0]) if isinstance(start_row['Close'], pd.Series) else start_row['Close'], 2)
            end_price = round(float(end_row['Close'].iloc[0]) if isinstance(end_row['Close'], pd.Series) else end_row['Close'], 2)
            change = round(end_price - start_price, 2)

            row[f"{year} JULY"] = start_price
            row[f"{year} AUG"] = end_price
            row[f"CHANGE {year}"] = change

        except:
            row[f"{year} JULY"] = "N/A"
            row[f"{year} AUG"] = "N/A"
            row[f"CHANGE {year}"] = "N/A"
    rows.append(row)

# 📊 Create DataFrame and export
df = pd.DataFrame(rows)
xlsx_file = "stock_july_aug_last5years_colored.xlsx"
df.to_excel(xlsx_file, index=False)

# 🎨 Add cell coloring based on change
wb = load_workbook(xlsx_file)
ws = wb.active

# Color codes
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
red = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

# Apply coloring to CHANGE columns
for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
    header = ws.cell(row=1, column=col[0].column).value
    if "CHANGE" in str(header):
        for cell in col:
            try:
                val = float(cell.value)
                if val > 50:
                    cell.fill = green
                elif val >= 0:
                    cell.fill = yellow
                else:
                    cell.fill = red
            except:
                continue  # Skip N/A or empty

wb.save(xlsx_file)
print(f"✅ Excel saved as: {xlsx_file}")
